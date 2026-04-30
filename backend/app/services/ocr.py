import io
import csv
import logging
import asyncio
from typing import Optional, List, Dict, Any

import boto3
from botocore.exceptions import ClientError, NoCredentialsError
from pdf2image import convert_from_bytes
import fitz  # PyMuPDF
from PIL import Image

from app.config import get_settings

logger = logging.getLogger(__name__)
settings = get_settings()


def _extract_from_excel(file_bytes: bytes, file_type: str) -> dict:
    """
    Extract text content from Excel (.xlsx, .xls) or CSV files.
    Converts spreadsheet data to structured plain text for AI ingestion.
    """
    text_parts = []

    if file_type == "csv":
        try:
            content = file_bytes.decode("utf-8", errors="replace")
            reader = csv.reader(io.StringIO(content))
            rows = list(reader)
            if rows:
                header = rows[0]
                text_parts.append("Columns: " + " | ".join(str(c) for c in header))
                text_parts.append("")
                for i, row in enumerate(rows[1:], 1):
                    if any(cell.strip() for cell in row):
                        row_text = " | ".join(str(c) for c in row)
                        text_parts.append(f"Row {i}: {row_text}")
            return {
                "text": "\n".join(text_parts),
                "confidence": 95.0,
                "blocks": [],
                "method": "csv-parse",
            }
        except Exception as e:
            logger.error(f"CSV parse failed: {e}")
            return {"text": "", "confidence": 0.0, "blocks": [], "method": "none"}

    # Excel: .xlsx or .xls
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text_parts.append(f"=== Sheet: {sheet_name} ===")
            rows_found = 0
            for row in ws.iter_rows(values_only=True):
                # Skip entirely empty rows
                if not any(cell is not None and str(cell).strip() for cell in row):
                    continue
                row_text = " | ".join(str(c) if c is not None else "" for c in row)
                text_parts.append(row_text)
                rows_found += 1
                if rows_found >= 500:  # Cap at 500 rows per sheet to avoid token overflow
                    text_parts.append(f"[Truncated after 500 rows]")
                    break
            text_parts.append("")
        wb.close()

        full_text = "\n".join(text_parts)
        return {
            "text": full_text,
            "confidence": 95.0,
            "blocks": [],
            "method": "openpyxl",
        }
    except Exception as e:
        logger.error(f"openpyxl extraction failed: {e}")
        # Fallback: try xlrd for older .xls files
        if file_type == "xls":
            try:
                import xlrd
                wb = xlrd.open_workbook(file_contents=file_bytes)
                for sheet_idx in range(wb.nsheets):
                    ws = wb.sheet_by_index(sheet_idx)
                    text_parts.append(f"=== Sheet: {ws.name} ===")
                    for row_idx in range(min(ws.nrows, 500)):
                        row_text = " | ".join(str(ws.cell_value(row_idx, col)) for col in range(ws.ncols))
                        text_parts.append(row_text)
                    text_parts.append("")
                return {
                    "text": "\n".join(text_parts),
                    "confidence": 90.0,
                    "blocks": [],
                    "method": "xlrd",
                }
            except Exception as e2:
                logger.error(f"xlrd fallback also failed: {e2}")

        return {"text": "", "confidence": 0.0, "blocks": [], "method": "none"}


async def extract_text_from_file(file_bytes: bytes, file_type: str) -> dict:
    """
    Extract text from a document using AWS Textract (PDFs/images) or
    native parsers for Excel/CSV.
    
    Returns:
        {
            "text": str,
            "confidence": float (0-100),
            "blocks": list[dict],
            "method": str
        }
    """
    # Excel and CSV: use native parsers (no OCR needed)
    if file_type in ("xlsx", "xls", "csv"):
        logger.info(f"OCR: Using native parser for {file_type} file")
        return await asyncio.to_thread(_extract_from_excel, file_bytes, file_type)

    # For institutional grade processing of PDFs, we convert to images first
    if file_type == "pdf":
        logger.info("OCR: Starting institutional-grade PDF-to-image preprocessing")
        try:
            return await _extract_via_image_rendering(file_bytes)
        except Exception as e:
            logger.error(f"Institutional OCR failed: {e}. Falling back to basic extraction.")
    
    # Try AWS Textract direct (for non-PDFs or as fallback)
    try:
        res = await _extract_with_textract(file_bytes)
        if len(res.get("text", "").strip()) > 100:
            return res
    except (NoCredentialsError, ClientError) as e:
        logger.warning(f"Textract primary extraction unavailable: {e}")

    # Ultimate fallback: local text extraction
    if file_type == "pdf":
        logger.info("Direct Textract/OCR failed, trying local PyPDF2 text extraction...")
        try:
            local_res = _extract_with_pypdf2(file_bytes)
            if len(local_res.get("text", "").strip()) > 10:
                return local_res
        except Exception as e:
            logger.error(f"PyPDF2 fallback failed: {e}")

    return {
        "text": "",
        "confidence": 0.0,
        "blocks": [],
        "method": "none",
    }


async def _extract_with_textract(file_bytes: bytes) -> dict:
    """Use AWS Textract to detect text in a document."""
    def _call_textract():
        client = boto3.client(
            "textract",
            region_name=settings.AWS_REGION,
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
        )
        return client.detect_document_text(
            Document={"Bytes": file_bytes}
        )

    response = await asyncio.to_thread(_call_textract)

    blocks = response.get("Blocks", [])
    lines = []
    confidences = []

    for block in blocks:
        if block["BlockType"] == "LINE":
            lines.append(block.get("Text", ""))
            confidences.append(block.get("Confidence", 0))

    text = "\n".join(lines)
    avg_confidence = sum(confidences) / len(confidences) if confidences else 0.0

    return {
        "text": text,
        "confidence": round(avg_confidence, 2),
        "blocks": [
            {
                "text": b.get("Text", ""),
                "confidence": b.get("Confidence", 0),
                "type": b.get("BlockType", ""),
                "geometry": b.get("Geometry", {}),
            }
            for b in blocks
            if b["BlockType"] in ("LINE", "WORD")
        ],
        "method": "textract",
    }


def _extract_with_pypdf2(file_bytes: bytes) -> dict:
    """Fallback: extract text from PDF using PyPDF2."""
    import io
    from PyPDF2 import PdfReader

    reader = PdfReader(io.BytesIO(file_bytes))
    pages_text = []

    for page in reader.pages:
        text = page.extract_text()
        if text:
            pages_text.append(text.strip())

    return {
        "text": "\n\n".join(pages_text),
        "confidence": 70.0,  # Arbitrary confidence for direct text extraction
        "blocks": [],
        "method": "pypdf2",
    }

    full_text = "\n\n".join(pages_text)

    return {
        "text": full_text,
        "confidence": 70.0 if full_text else 0.0,  # Moderate confidence for PDF extraction
        "blocks": [],
        "method": "pypdf2",
    }


async def _extract_via_image_rendering(file_bytes: bytes) -> dict:
    """
    Institutional Grade PDF Extraction:
    1. Renders PDF pages to high-DPI images (300 DPI)
    2. Runs Textract on each image
    3. Merges results with page tracking
    """
    # 1. Convert PDF to images
    img_byte_arrays = await _convert_pdf_to_images(file_bytes)
    
    all_text = []
    all_blocks = []
    all_confidences = []

    logger.info(f"OCR: Processing {len(img_byte_arrays)} rendered pages")

    for i, img_bytes in enumerate(img_byte_arrays):
        # OCR the image
        res = await _extract_with_textract(img_bytes)
        
        if res["text"]:
            # Add page marker for RAG and clarity
            page_text = f"--- Page {i+1} ---\n{res['text']}"
            all_text.append(page_text)
            
            # Update block metadata with page number
            for block in res["blocks"]:
                block["page"] = i + 1
            
            all_blocks.extend(res["blocks"])
            all_confidences.append(res["confidence"])
            logger.info(f"OCR: Page {i+1} completed ({len(res['text'])} chars)")
        else:
            logger.warning(f"OCR: Page {i+1} returned no text")

    full_text = "\n\n".join(all_text)
    avg_confidence = sum(all_confidences) / len(all_confidences) if all_confidences else 0.0

    return {
        "text": full_text,
        "confidence": round(avg_confidence, 2),
        "blocks": all_blocks,
        "method": "textract-institutional-image",
    }


async def _convert_pdf_to_images(file_bytes: bytes) -> List[bytes]:
    """
    Preprocessing step: Converts PDF pages to high-DPI images.
    Uses pdf2image (poppler) if available, falls back to PyMuPDF (fitz).
    """
    try:
        # Try pdf2image (institutional preference)
        images = await asyncio.to_thread(convert_from_bytes, file_bytes, dpi=300)
        img_byte_arrays = []
        for img in images:
            img_byte_arr = io.BytesIO()
            img.save(img_byte_arr, format='PNG')
            img_byte_arrays.append(img_byte_arr.getvalue())
        logger.info("PDF Preprocessing: Successfully converted via pdf2image")
        return img_byte_arrays
    except Exception as e:
        logger.warning(f"pdf2image conversion failed or poppler not found: {e}. Using PyMuPDF fallback.")
        
        # Fallback to PyMuPDF (fitz) - very reliable and usually already installed
        try:
            doc = fitz.open(stream=file_bytes, filetype="pdf")
            img_byte_arrays = []
            for page in doc:
                # Zoom 2x = 144 DPI, Zoom 4.16x = 300 DPI approx (standard is 72 DPI)
                # Matrix(2, 2) is ~150 DPI. For 300 DPI we want ~4.16
                zoom = 300 / 72
                mat = fitz.Matrix(zoom, zoom)
                pix = page.get_pixmap(matrix=mat)
                img_byte_arrays.append(pix.tobytes("png"))
            doc.close()
            logger.info("PDF Preprocessing: Successfully converted via PyMuPDF")
            return img_byte_arrays
        except Exception as e2:
            logger.error(f"All PDF-to-image conversion methods failed: {e2}")
            raise
